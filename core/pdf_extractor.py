# -*- coding: utf-8 -*-
"""
pdf_extractor.py — PDF 불량보고서 → Excel 변환 (GUI 없는 서버용 버전)
원본: bulyang_rate_analyzer_v9.py (created by 김지연)
의존성: pymupdf openpyxl
"""

import os
import re
from datetime import datetime

import fitz  # PyMuPDF
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT_SHEET_RAW    = "① 원본백데이터"
OUTPUT_SHEET_DETAIL = "② 불량상세"
OUTPUT_SHEET_SUMMARY = "③ 요약"

RAW_HEADERS = [
    "파일명", "지역1", "지역2", "INS. DATE_시작일", "REPORT NO.", "이전 보고서번호",
    "바이어_수정", "의뢰업체", "브랜드", "공장",
    "스타일번호", "품명", "검사종류", "ORDER Q'TY", "INSPEC. Q'TY", "PASS Q'TY", "FAIL Q'TY",
    "1차검사수량", "1차합격수량", "1차불합격수량", "2차검사수량", "2차합격수량", "최종불합격수량",
]
for i in range(1, 21):   # 19 → 20으로 확장
    RAW_HEADERS += [f"주요불량{i}", f"불량갯수{i}"]

DETAIL_HEADERS = [
    "파일명", "REPORT NO.", "이전 보고서번호", "검사일", "바이어", "의뢰업체", "브랜드",
    "공장", "지역1", "지역2", "스타일번호", "품명", "검사수량(INSPEC)",
    "1차불합격수량", "최종불합격수량", "2차검사수량",
    "원본불량명", "중불량", "경불량", "불량수량"
]


# ── 유틸 ─────────────────────────────────────────────────────────
def normalize_space(s):
    s = str(s or "")
    s = s.replace("￾", " ").replace("﻿", " ").replace("\x01", " ").replace("\x00", " ").replace(" ", " ")
    return re.sub(r"\s+", " ", s).strip()

def compact_label(s):
    return re.sub(r"[\s ￾﻿\x00-\x1f]+", "", str(s or ""))

def to_int(value):
    if value is None:
        return None
    s = str(value).replace(",", "").replace("PCS", "").replace("pcs", "").strip()
    m = re.search(r"-?\d+", s)
    return int(m.group()) if m else None

def first_match(pattern, text, flags=re.S):
    m = re.search(pattern, text, flags)
    return normalize_space(m.group(1)) if m else None


# ── 날짜 추출 ──────────────────────────────────────────────────────
def extract_ins_start_date(text):
    text = normalize_space(text)
    windows = []
    for m in re.finditer(r"INS\.?\s*DATE", text, re.I):
        windows.append(text[m.end():m.end() + 80])
    windows.append(text)
    patterns = [
        r"([12][0-9]{3})\s*[.\-/]\s*([0-9]{1,2})\s*[.\-/]\s*([0-9]{1,2})",
        r"\b([12][0-9]{3})([01][0-9])([0-3][0-9])\b",
    ]
    for win in windows:
        for pat in patterns:
            m = re.search(pat, win)
            if m:
                y, mth, d = m.groups()
                return f"{int(y):04d}.{int(mth):02d}.{int(d):02d}"
    return None


# ── PDF 텍스트/테이블 추출 ──────────────────────────────────────────
def text_from_pdf(doc):
    return normalize_space("\n".join(page.get_text("text") for page in doc))

def find_value_in_table(table_rows, label):
    target = compact_label(label)
    for row in table_rows or []:
        row = [normalize_space(c) for c in row]
        compact = [compact_label(c) for c in row]
        for i, c in enumerate(compact):
            if c == target and i + 1 < len(row):
                value = normalize_space(row[i + 1])
                if value:
                    return value
    for r, row in enumerate(table_rows or []):
        compact = [compact_label(c) for c in row]
        for i, c in enumerate(compact):
            if c == target and r + 1 < len(table_rows) and i < len(table_rows[r + 1]):
                value = normalize_space(table_rows[r + 1][i])
                if value:
                    return value
    return None

def find_header_text(table_rows):
    for row in table_rows:
        for c in row:
            c = normalize_space(c)
            if "제품 평가 보고서" in c and "REPORT NO" in c:
                return c
    return ""

def extract_tables(page):
    try:
        return page.find_tables().tables
    except Exception:
        return []


# ── 합계 행 정규화 ──────────────────────────────────────────────────
def _filter_color_total(color_total):
    """
    전수평가 합계 행에서 빈 셀과 퍼센트 값을 제거합니다.
    PDF 포맷에 따라 빈 열 수가 달라 고정 인덱스가 어긋나는 문제를 해결합니다.
    반환: ['합계', 1차검사수량, 1차불합격수량, 최종불합격수량,
            재평가합격수량, 최종합격수량(PASS), 총검사수량(INSPEC)]
    """
    _pct = re.compile(r"^\d+\.?\d*\s*%$")
    return ["합계"] + [
        x for x in color_total[1:]
        if x and not _pct.match(x.replace(",", "").strip())
    ]


# ── 불량 내용 파싱 ──────────────────────────────────────────────────
def parse_defects(table_rows, full_text):
    """
    '주요 불량 내용' 표를 정확히 추출합니다.

    v9 수정사항:
    - 헤더 행을 동적으로 탐지해 열 위치 결정 (단일·2-table 포맷 모두 지원)
    - table0 + table1 통합 전달 필요 (호출부에서 처리)
    - 합 계 행과 개별 합 비교 → 차이를 '기타'로 보정 (PDF 추출 누락 대응)
    - 불량 항목 상한 19 → 20
    """
    rows = [[normalize_space(c) for c in (row or [])] for row in (table_rows or [])]

    defects = []
    header_idx = None
    name_col = major_col = minor_col = None

    # 1) '주요 불량 내용 / 중불량 / 경불량' 헤더 행 탐지
    for i, row in enumerate(rows):
        compact = [compact_label(c) for c in row]
        if (any("주요불량내용" in c for c in compact)
                and any("중불량" in c for c in compact)
                and any("경불량" in c for c in compact)):
            header_idx = i
            for j, c in enumerate(compact):
                if name_col is None and "주요불량내용" in c:
                    name_col = j
                if major_col is None and "중불량" in c:
                    major_col = j
                if minor_col is None and "경불량" in c:
                    minor_col = j
            break

    total_major_from_sum = None
    total_minor_from_sum = None

    # 2) 헤더 아래 행 읽기, 합계/다음 섹션에서 종료
    if header_idx is not None and name_col is not None and major_col is not None and minor_col is not None:
        for row in rows[header_idx + 1:]:
            if not row:
                continue
            first = compact_label(row[0] if row else "")
            joined = "".join(compact_label(c) for c in row)

            if first in ("합계", "합계계") or first.startswith("4.확인사항") or "4.확인사항" in joined:
                if first in ("합계", "합계계"):
                    total_major_from_sum = to_int(row[major_col] if major_col < len(row) else None)
                    total_minor_from_sum = to_int(row[minor_col] if minor_col < len(row) else None)
                break

            name = normalize_space(row[name_col] if name_col < len(row) else "")
            if not name:
                continue
            if name in ("주요 불량 내용", "중불량", "경불량", "불량 발견 부위", "부위참고"):
                continue
            if name.startswith("-") or re.match(r"^\d+\.", name):
                continue

            major = to_int(row[major_col] if major_col < len(row) else None) or 0
            minor = to_int(row[minor_col] if minor_col < len(row) else None) or 0
            if major + minor > 0:
                defects.append({"name": name, "major": major, "minor": minor, "qty": major + minor})

        if defects:
            # 합 계 행과 개별 합 비교 → 차이가 있으면 '기타'로 보정
            if total_major_from_sum is not None or total_minor_from_sum is not None:
                diff_major = (total_major_from_sum or 0) - sum(d["major"] for d in defects)
                diff_minor = (total_minor_from_sum or 0) - sum(d["minor"] for d in defects)
                if diff_major > 0 or diff_minor > 0:
                    defects.append({
                        "name": "기타",
                        "major": max(0, diff_major),
                        "minor": max(0, diff_minor),
                        "qty": max(0, diff_major) + max(0, diff_minor)
                    })
            return defects[:20]

    # 3) 구형/다른 양식 fallback
    section_rows = []
    in_section = False
    for row in rows:
        joined = "".join(compact_label(c) for c in row)
        first = compact_label(row[0] if row else "")
        if "주요불량내용" in joined and (
            first.startswith("3.") or first.startswith("4.") or first == "주요불량내용"
        ):
            in_section = True
            continue
        if in_section and ("확인사항" in joined or first in ("합계", "합계계")):
            break
        if in_section:
            section_rows.append(row)

    for row in section_rows:
        if not row:
            continue
        name = normalize_space(row[0] if len(row) > 0 else "")
        if not name or name.startswith("-") or re.match(r"^\d+\.", name):
            continue
        major = to_int(row[1] if len(row) > 1 else None) or 0
        minor = to_int(row[2] if len(row) > 2 else None) or 0
        if major + minor > 0:
            defects.append({"name": name, "major": major, "minor": minor, "qty": major + minor})

    if defects:
        return defects[:20]

    # 4) 최후 fallback: 텍스트 기반 추출
    defect_names = [
        "바텍 누락", "좌우 비대칭", "제사처리불량", "제사처리 미흡",
        "원단불량", "원단 불량", "기름오염", "구멍", "이색", "염반",
        "봉탈", "봉비", "퍼커링", "히까리", "잡사", "오염",
        "찝힘", "바늘자국", "늘어짐", "찢어짐"
    ]
    text = normalize_space(full_text)
    m = re.search(r"(?:3|4)\.\s*주요\s*불량\s*내용(.*?)(?:4|5)\.\s*확인사항", text, re.S)
    defect_text = m.group(1) if m else ""
    for name in defect_names:
        pat = re.compile(re.escape(name) + r"\s+(\d+)(?:\s+(\d+))?")
        mm = pat.search(defect_text)
        if mm:
            major = to_int(mm.group(1)) or 0
            minor = to_int(mm.group(2)) or 0
            defects.append({"name": name, "major": major, "minor": minor, "qty": major + minor})
    return defects[:20]


# ── 파일명 메타데이터 추출 ───────────────────────────────────────────
def split_filename_metadata(filename):
    base = os.path.splitext(os.path.basename(filename))[0]
    base = re.sub(r"^\(양식변경\)\s*", "", base).strip()
    report = first_match(r"([A-Z][0-9]{3}-[0-9]{2}-[0-9]{5}(?:-[0-9]{2})?)", base, flags=0)
    meta = {"REPORT NO.": report}
    if not report:
        return meta
    tail = base.split(report, 1)[-1].lstrip("_ -")
    parts = [p for p in tail.split("_") if p]
    if len(parts) >= 1:
        meta["의뢰업체"] = parts[0]
    if len(parts) >= 2:
        meta["브랜드"] = parts[1].replace("BEBEDEPINO", "BEBE DE PINO").replace("HENRY_COTTON_S", "HENRY COTTON'S")
    code_candidates = []
    for p in parts[2:]:
        if re.search(r"[A-Za-z]", p) and re.search(r"\d", p) and len(p) >= 6:
            if not re.search(r"^(?:1차|2차|3차|전수|샘플|추가|최초|최조|\d{8})", p):
                code_candidates.append(p)
    if code_candidates:
        meta["스타일번호"] = code_candidates[0]
        try:
            idx = parts.index(code_candidates[0])
            if idx > 2:
                meta["품명"] = " ".join(parts[2:idx])
        except ValueError:
            pass
    m = re.search(r"(20\d{2})(\d{2})(\d{2})", base)
    if m:
        y, mo, d = m.groups()
        meta["INS. DATE_시작일"] = f"{y}.{mo}.{d}"
    return {k: normalize_space(v) for k, v in meta.items() if v}

def apply_filename_fallback(rec, filename):
    meta = split_filename_metadata(filename)
    for k, v in meta.items():
        if rec.get(k) in (None, ""):
            rec[k] = v

def normalize_country_to_region(country):
    country = normalize_space(country).upper()
    _map = {"VIETNAM": "베트남", "VIET": "베트남", "VIET NAM": "베트남",
            "KOREA": "한국", "CHINA": "중국", "MYANMAR": "미얀마"}
    return _map.get(country, country or None)

def normalize_local_region(region):
    region = normalize_space(region)
    return {"HCMC": "호치민", "DANDONG": "단동"}.get(region, region or None)

def is_bad_table_value(value):
    value = normalize_space(value)
    if not value:
        return True
    bad = ["공장 평가상태", "평가 준비 상태", "사전 완성반평가", "신청업체", "벤더 참관"]
    return any(b in value for b in bad)

def fallback_basic_fields_from_text(rec, full_text, filename):
    text = normalize_space(full_text)
    if not rec.get("REPORT NO."):
        rec["REPORT NO."] = first_match(r"REPORT\s*NO\.?\s*([A-Z0-9\-]+)", text)
    if not rec.get("INS. DATE_시작일"):
        rec["INS. DATE_시작일"] = extract_ins_start_date(text)
    if not rec.get("지역1"):
        rec["지역1"] = normalize_country_to_region(first_match(r"COUNTRY\s+([A-Z]+)", text) or "")
    if not rec.get("스타일번호"):
        meta = split_filename_metadata(filename)
        rec["스타일번호"] = meta.get("스타일번호") or first_match(r"품\s*번\s*([A-Z0-9\-]+)", text)
    m_qty = re.search(r"([0-9,]+)\s*(?:PCS|pcs)\s+([0-9,]+)\s*(?:PCS|pcs)\s+([^\n]*?평가)", full_text, re.I)
    if m_qty:
        if rec.get("ORDER Q'TY") in (None, ""):
            rec["ORDER Q'TY"] = to_int(m_qty.group(1))
        if rec.get("INSPEC. Q'TY") in (None, ""):
            rec["INSPEC. Q'TY"] = to_int(m_qty.group(2))
        if not rec.get("검사종류"):
            rec["검사종류"] = normalize_space(m_qty.group(3))
    apply_filename_fallback(rec, filename)
    return rec


# ── 메인 PDF 파싱 함수 ─────────────────────────────────────────────
def parse_pdf(pdf_path: str) -> dict:
    """PDF 한 장 파싱 → record dict 반환. 실패 시 예외 raise."""
    filename = os.path.basename(pdf_path)
    doc = fitz.open(pdf_path)
    try:
        if len(doc) == 0:
            raise ValueError("빈 PDF")

        full_text = text_from_pdf(doc)
        is_full_inspection = "색상별 제품평가 수량" in full_text or "전수평가" in full_text
        is_sampling = "색상별 오더수량" in full_text or "샘플링 평가" in full_text or "Sampling Plan" in full_text

        if "제품 평가 보고서" not in full_text:
            raise ValueError("문서 제목 확인 실패: '제품 평가 보고서' 텍스트를 찾을 수 없음")
        if not (is_full_inspection or is_sampling):
            raise ValueError("평가 방식 확인 실패: 전수/샘플링 방식을 식별할 수 없음")

        tables = extract_tables(doc[0])
        table0 = tables[0].extract() if len(tables) >= 1 else []
        table1 = tables[1].extract() if len(tables) >= 2 else []
        header_text = find_header_text(table0) or full_text

        rec = {"파일명": filename}
        rec["REPORT NO."] = (first_match(r"REPORT\s*NO\.?\s*([A-Z0-9\-]+)", header_text)
                             or first_match(r"REPORT\s*NO\.?\s*([A-Z0-9\-]+)", full_text)
                             or first_match(r"\(([A-Z0-9\-]+)\)", filename))
        rec["INS. DATE_시작일"] = extract_ins_start_date(header_text) or extract_ins_start_date(full_text)
        country = first_match(r"COUNTRY\s+([A-Z가-힣]+)", header_text) or first_match(r"COUNTRY\s+([A-Z가-힣]+)", full_text)
        rec["지역1"] = normalize_country_to_region(country)

        rec["의뢰업체"] = find_value_in_table(table0, "의뢰업체")
        prev_report_no = find_value_in_table(table0, "이전 보고서번호")
        rec["이전 보고서번호"] = prev_report_no if prev_report_no and prev_report_no.strip() not in ("-", "–", "—") else ""
        rec["바이어_수정"] = find_value_in_table(table0, "바이어")
        rec["공장"] = find_value_in_table(table0, "공장")
        if is_bad_table_value(rec.get("공장")):
            rec["공장"] = None
        rec["브랜드"] = find_value_in_table(table0, "브랜드")
        region_value = find_value_in_table(table0, "지역")
        rec["지역2"] = normalize_local_region(region_value) if region_value and not re.match(r"^[A-Z][0-9]{3}-", region_value) else None
        rec["품명"] = find_value_in_table(table0, "품 명")
        rec["스타일번호"] = find_value_in_table(table0, "품 번") or first_match(r"([A-Z]{3,}\d+[A-Z0-9]*)", filename)
        rec["검사종류"] = find_value_in_table(table0, "평가종류")
        rec["ORDER Q'TY"] = to_int(find_value_in_table(table0, "총오더수량"))

        m_qty = re.search(r"([0-9,]+)\s*PCS\s+([0-9,]+)\s*PCS\s+([^\n]*?평가)", full_text)
        second_qty = None
        if m_qty:
            rec["ORDER Q'TY"] = rec.get("ORDER Q'TY") or to_int(m_qty.group(1))
            rec["검사종류"] = rec.get("검사종류") or normalize_space(m_qty.group(3))
            second_qty = to_int(m_qty.group(2))
        if not rec.get("검사종류"):
            rec["검사종류"] = "샘플링 평가" if is_sampling else "전수평가"

        fallback_basic_fields_from_text(rec, full_text, filename)

        if is_full_inspection and not is_sampling:
            color_total = None
            for row in table0:
                row = [normalize_space(c) for c in row]
                if row and row[0] == "합계":
                    color_total = row
                    break
            if not color_total:
                m = re.search(r"합계\s+([0-9,]+)\s+([0-9,]+)\s+(\d+)\s+(\d+)\s+([0-9,]+)\s+([0-9,]+)\s+[0-9.]+%", full_text)
                if m:
                    color_total = ["합계", "", m.group(1), m.group(2), "", "", m.group(3), "", m.group(4), m.group(5), "", m.group(6)]
            if not color_total:
                raise ValueError("색상별 제품평가 수량 합계 행을 찾을 수 없음")

            # 빈 셀·% 값 제거 후 인덱스 적용 (포맷별 빈 열 수 차이 대응)
            ct = _filter_color_total(color_total)
            rec["1차검사수량"] = to_int(ct[1] if len(ct) > 1 else None) or second_qty
            rec["1차불합격수량"] = to_int(ct[2] if len(ct) > 2 else None)
            rec["최종불합격수량"] = to_int(ct[3] if len(ct) > 3 else None)
            rec["2차검사수량"] = rec["1차불합격수량"]
            rec["2차합격수량"] = to_int(ct[4] if len(ct) > 4 else None)
            rec["PASS Q'TY"] = to_int(ct[5] if len(ct) > 5 else None)
            rec["FAIL Q'TY"] = rec["최종불합격수량"]
            rec["INSPEC. Q'TY"] = to_int(ct[6] if len(ct) > 6 else None) or rec["1차검사수량"]
            rec["1차합격수량"] = (rec.get("1차검사수량") or 0) - (rec.get("1차불합격수량") or 0)
        else:
            shipment_total = None
            sample_total = None
            for row in table0:
                row = [normalize_space(c) for c in row]
                if row and row[0] == "합계":
                    nums = [to_int(x) for x in row if to_int(x) is not None]
                    if len(nums) >= 2:
                        shipment_total, sample_total = nums[0], nums[1]
                    break
            if shipment_total is None or sample_total is None:
                m = re.search(r"합계\s+([0-9,]+)\s+([0-9,]+)\s+<\s*합\s*격\s*>", full_text)
                if m:
                    shipment_total = to_int(m.group(1))
                    sample_total = to_int(m.group(2))

            rec["1차검사수량"] = shipment_total or second_qty
            rec["INSPEC. Q'TY"] = sample_total or second_qty
            found = re.search(r"불량발견매수\s+(\d+)\s+(\d+)", full_text)
            rec["1차불합격수량"] = (to_int(found.group(1)) or 0) + (to_int(found.group(2)) or 0) if found else None
            rec["최종불합격수량"] = 0 if ("<합 격>" in full_text or "<합격>" in full_text or "합격허용매수 이내" in full_text) else rec.get("1차불합격수량")
            rec["FAIL Q'TY"] = rec["최종불합격수량"]
            rec["PASS Q'TY"] = (rec.get("INSPEC. Q'TY") or 0) - (rec.get("FAIL Q'TY") or 0) if rec.get("INSPEC. Q'TY") is not None else None
            rec["1차합격수량"] = (rec.get("INSPEC. Q'TY") or 0) - (rec.get("1차불합격수량") or 0) if rec.get("INSPEC. Q'TY") is not None and rec.get("1차불합격수량") is not None else None
            rec["2차검사수량"] = None
            rec["2차합격수량"] = None

        # table0 + table1 통합 전달 (단일 테이블 PDF 대응)
        rec["defects"] = parse_defects((table0 or []) + (table1 or []), full_text)

        required = ["REPORT NO.", "INS. DATE_시작일", "스타일번호", "INSPEC. Q'TY"]
        missing = [k for k in required if rec.get(k) in (None, "")]
        if missing:
            raise ValueError("필수 필드 누락: " + ", ".join(missing))
        return rec
    finally:
        doc.close()


# ── Excel 생성 ────────────────────────────────────────────────────
def make_workbook(records: list[dict], output_path: str):
    """records 리스트로 Excel 파일 생성 후 output_path에 저장."""
    wb = Workbook()
    ws = wb.active
    ws.title = OUTPUT_SHEET_RAW
    detail  = wb.create_sheet(OUTPUT_SHEET_DETAIL)
    summary = wb.create_sheet(OUTPUT_SHEET_SUMMARY)

    title_fill  = PatternFill("solid", fgColor="E2F0D9")
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    thin   = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.append(["▶ 원본 백데이터 PDF 자동 추출"])
    ws.append(RAW_HEADERS)
    for rec in records:
        base_col_count = 23
        row = [rec.get(h) for h in RAW_HEADERS[:base_col_count]]
        defect_values = []
        for d in rec.get("defects", []):
            defect_values += [d["name"], d["qty"]]
        while len(defect_values) < 40:   # 20개 × 2열
            defect_values += [None, None]
        ws.append(row + defect_values[:40])

    detail.append(["▶ 불량항목 상세"])
    detail.append(DETAIL_HEADERS)
    for rec in records:
        for d in rec.get("defects", []):
            detail.append([
                rec.get("파일명"), rec.get("REPORT NO."), rec.get("이전 보고서번호"),
                rec.get("INS. DATE_시작일"), rec.get("바이어_수정"), rec.get("의뢰업체"),
                rec.get("브랜드"), rec.get("공장"), rec.get("지역1"), rec.get("지역2"),
                rec.get("스타일번호"), rec.get("품명"),
                rec.get("INSPEC. Q'TY"),
                rec.get("1차불합격수량"), rec.get("최종불합격수량"), rec.get("2차검사수량"),
                d["name"], d["major"], d["minor"], d["qty"]
            ])

    summary.append(["▶ PDF 통합 요약"])
    summary.append(["생성일시", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    summary.append(["총 보고서 수", len(records)])
    summary.append(["총 검사 수량", sum((r.get("INSPEC. Q'TY") or 0) for r in records)])
    summary.append(["총 1차불량수량", sum((r.get("1차불합격수량") or 0) for r in records)])
    summary.append(["총 최종불합격수량", sum((r.get("최종불합격수량") or 0) for r in records)])

    for sheet in [ws, detail, summary]:
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                cell.border = border
                if cell.row == 1:
                    cell.font = Font(bold=True, size=12)
                    cell.fill = title_fill
                elif cell.row == 2:
                    cell.font = Font(bold=True)
                    cell.fill = header_fill
        for col in range(1, sheet.max_column + 1):
            max_len = 0
            for cell in sheet[get_column_letter(col)]:
                max_len = max(max_len, len(str(cell.value or "")))
            sheet.column_dimensions[get_column_letter(col)].width = min(max(max_len + 2, 10), 32)
        sheet.freeze_panes = "A3"

    wb.save(output_path)
