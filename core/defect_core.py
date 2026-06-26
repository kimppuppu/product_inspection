"""
defect_core.py — 불량명 표준화 매핑 핵심 함수 모듈
불량명_매핑_프로그램.py v3.4 에서 GUI(tkinter) 제거 후 API용으로 추출
"""
import os, re
from collections import defaultdict

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.cell.rich_text import CellRichText, TextBlock
    from openpyxl.cell.text import InlineFont
    from rapidfuzz import fuzz
except ImportError:
    import subprocess, sys
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "rapidfuzz"])
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.cell.rich_text import CellRichText, TextBlock
    from openpyxl.cell.text import InlineFont
    from rapidfuzz import fuzz

TH_HIGH = 85
TH_LOW  = 65

# 최근 업데이트(별칭 추가)된 내용을 표시할 빨간색 글자
RED_FONT = InlineFont(color="FFFF0000")


def _rich_text_to_str(val) -> str:
    """셀 값이 서식 적용된 리치텍스트(CellRichText)여도 일반 문자열로 변환."""
    if val is None:
        return ""
    if isinstance(val, CellRichText):
        parts = []
        for p in val:
            parts.append(p.text if isinstance(p, TextBlock) else str(p))
        return "".join(parts)
    return str(val)


# ── 표준불량명칭 시트 자동 감지 ─────────────────────────────────
def find_std_sheet(wb):
    def has_defect_header(ws):
        for row in ws.iter_rows(min_row=1, max_row=3, values_only=True):
            for cell in row:
                if cell and '불량명' in str(cell):
                    return True
        return False

    if '최종본' in wb.sheetnames:
        return wb['최종본']
    if '최종' in wb.sheetnames:
        return wb['최종']
    if '2차수정본' in wb.sheetnames:
        return wb['2차수정본']
    for n in wb.sheetnames:
        if ('최종' in n or '불량명' in n or '표준' in n or '수정' in n) and has_defect_header(wb[n]):
            return wb[n]
    for n in wb.sheetnames:
        if n not in ('부위구분',) and has_defect_header(wb[n]):
            return wb[n]
    for n in wb.sheetnames:
        if n not in ('부위구분',):
            return wb[n]
    return wb[wb.sheetnames[0]]


# ── 표준불량명칭 로드 ────────────────────────────────────────────
def load_standard(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = find_std_sheet(wb)
    used = ws.title
    names, adict, cat = [], {}, None

    for row in ws.iter_rows(min_row=2, values_only=True):
        c, no, name = row[0], row[1], row[2]
        desc = row[6] if len(row) > 6 else None
        desc_str = _rich_text_to_str(desc)
        if c: cat = c
        if not name: continue
        name = str(name).strip()
        if not name: continue
        names.append((name, cat, desc_str))
        adict[name] = name
        if desc_str:
            for cp in re.split(r'[,\n]', desc_str):
                for p in cp.split('/'):
                    p = p.strip()
                    if p and len(p) >= 2:
                        adict[p] = name

    if len(names) < 5:
        raise ValueError(f"표준불량명칭 데이터를 찾을 수 없습니다. 파일: {os.path.basename(path)}, 시트: '{used}'")
    if len(names) > 200:
        raise ValueError(f"선택한 파일이 표준불량명칭 파일이 아닌 것 같습니다. (항목 수: {len(names)}개, 정상: 약 76개)")
    return names, adict, used


# ── 헤더에서 컬럼 인덱스 매핑 ────────────────────────────────────
def parse_header(ws):
    COL_NAMES = {
        '파일명': 'file', 'REPORT NO.': 'report_no', '검사일': 'date',
        '바이어': 'buyer', '의뢰업체': 'client', '브랜드': 'brand',
        '공장': 'factory', '지역1': 'region1', '지역2': 'region2',
        '스타일번호': 'style', '품명': 'item', '검사수량(INSPEC)': 'inspec',
        '원본불량명': 'defect_raw', '중불량': 'qty_mid',
        '경불량': 'qty_light', '불량수량': 'qty_total',
    }
    for ri, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), 1):
        if not any(row):
            continue
        col_map = {}
        for ci, v in enumerate(row):
            if v and str(v).strip() in COL_NAMES:
                col_map[COL_NAMES[str(v).strip()]] = ci
        if 'defect_raw' in col_map:
            return col_map, ri + 1
    return {
        'file': 0, 'report_no': 1, 'date': 2, 'buyer': 3,
        'client': 4, 'brand': 5, 'factory': 6, 'region1': 7,
        'region2': 8, 'style': 9, 'item': 10, 'inspec': 11,
        'defect_raw': 12, 'qty_mid': 13, 'qty_light': 14, 'qty_total': 15,
    }, 3


# ── 데이터 파일 로드 ──────────────────────────────────────────────
def load_raw(file_paths, log_fn=None):
    TARGET_SHEET = '② 불량상세'
    rows = []
    skipped = []

    for fp in file_paths:
        fn = os.path.basename(fp)
        if fn.startswith('~$'):
            skipped.append(fn + "  ← Excel 임시파일")
            continue
        if log_fn: log_fn(f"  [{fn}] 읽는 중...")

        try:
            wb = openpyxl.load_workbook(fp, data_only=True)
        except Exception as e:
            skipped.append(fn + f"  ← 읽기 오류: {e}")
            continue

        snames = wb.sheetnames
        ws = None
        if TARGET_SHEET in snames:
            ws = wb[TARGET_SHEET]
        else:
            for sn in snames:
                if '불량상세' in sn and not sn.startswith('④'):
                    ws = wb[sn]
                    if log_fn: log_fn(f"    ※ '{sn}' 시트 사용")
                    break

        if ws is None:
            skipped.append(fn + f"  ← '② 불량상세' 시트 없음")
            continue

        col_map, data_start = parse_header(ws)
        dc = col_map.get('defect_raw', 12)

        def gv(row, key, default=None):
            idx = col_map.get(key)
            if idx is None: return default
            return row[idx] if len(row) > idx else default

        cnt = 0
        for row in ws.iter_rows(min_row=data_start, values_only=True):
            d = row[dc] if len(row) > dc else None
            if not d: continue
            ds = str(d).strip()
            if not re.search(r'[가-힣a-zA-Z]', ds): continue

            # 날짜 직렬화
            date_val = gv(row, 'date')
            if hasattr(date_val, 'strftime'):
                date_val = date_val.strftime('%Y-%m-%d')
            elif date_val:
                date_val = str(date_val)

            rows.append({
                'file': fn, 'report_no': gv(row, 'report_no'),
                'date': date_val, 'buyer': gv(row, 'buyer'),
                'client': gv(row, 'client'), 'brand': gv(row, 'brand'),
                'factory': str(gv(row, 'factory') or '').strip(),
                'region1': str(gv(row, 'region1') or '').strip(),
                'region2': str(gv(row, 'region2') or '').strip(),
                'style': gv(row, 'style'), 'item': gv(row, 'item'),
                'inspec': _to_int(gv(row, 'inspec')),
                'defect_raw': ds,
                'qty_mid': _to_int(gv(row, 'qty_mid')),
                'qty_light': _to_int(gv(row, 'qty_light')),
                'qty_total': _to_int(gv(row, 'qty_total')),
            })
            rows[-1]['region_label'] = get_region_label(rows[-1])
            cnt += 1

        if log_fn: log_fn(f"    → {cnt}행 로드 완료")
        if cnt == 0 and log_fn: log_fn("    !! 데이터 없음")

    return rows, skipped


def get_region_label(row: dict) -> str:
    """
    접수번호 앞자리 + 기존 지역 정보로 세분화 지역명 반환
    S → 중국 상해 / Q → 중국 청도 / Y → 중국 연태
    A → 인도네시아
    V → 베트남 호치민 or 베트남 하노이 (지역2 기준)
    H → 한국 or 미얀마 (지역1 기준)
    """
    rno    = str(row.get('report_no') or '').strip().upper()
    prefix = rno[0] if rno else ''
    r1     = str(row.get('region1') or '').strip()
    r2     = str(row.get('region2') or '').strip()

    if prefix == 'S':
        return '중국 상해'
    elif prefix == 'Q':
        return '중국 청도'
    elif prefix == 'Y':
        return '중국 연태'
    elif prefix == 'A':
        return '인도네시아'
    elif prefix == 'V':
        r2l = r2.lower()
        if '하노이' in r2l or 'hanoi' in r2l or 'ha noi' in r2l:
            return '베트남 하노이'
        return '베트남 호치민'
    elif prefix == 'H':
        r1l = r1.lower()
        if '미얀마' in r1l or 'myanmar' in r1l or 'myan' in r1l:
            return '미얀마'
        return '한국'
    else:
        return r1 or '기타'


def _to_int(val):
    try: return int(val or 0)
    except: return 0


# ── 수동 수정 내용을 표준불량명칭.xlsx에 저장 ─────────────────────
def save_corrections_to_std(std_path: str, corrections: list[dict], log_fn=None) -> int:
    """
    corrections: [{"part": "봉재불량", "std": "봉제불량"}, ...]
    표준불량명칭.xlsx 의 해당 표준명 설명 컬럼에 별칭으로 추가.
    반환: 추가된 별칭 수
    """
    if not corrections:
        return 0

    wb = openpyxl.load_workbook(std_path)
    ws = find_std_sheet(wb)

    # 표준명 → 설명 셀 매핑
    desc_cells = {}
    for row in ws.iter_rows(min_row=2):
        name_cell = row[2]
        if not name_cell.value:
            continue
        name = str(name_cell.value).strip()
        if len(row) > 6:
            desc_cells[name] = row[6]

    # 표준명별로 새로 추가할 별칭을 모은다
    new_aliases = defaultdict(list)
    for corr in corrections:
        part = str(corr.get("part", "")).strip()
        std  = str(corr.get("std",  "")).strip()
        if not part or not std or part == std:
            continue
        if std not in desc_cells:
            if log_fn: log_fn(f"  ⚠️ '{std}' 표준명을 찾을 수 없음, 건너뜀")
            continue

        cell = desc_cells[std]
        current = _rich_text_to_str(cell.value)
        # 이미 포함된 별칭인지 확인
        existing = {re.sub(r'\s+', '', x).lower()
                    for x in re.split(r'[,/\n]', current) if x.strip()}
        if re.sub(r'\s+', '', part).lower() in existing:
            continue  # 이미 있음
        new_aliases[std].append(part)

    # 표준명별로 한 번에 반영 — 새로 추가된 부분만 빨간색으로 표시
    # (이전에 추가되어 빨간색이었던 내용은 이번 업데이트 기준으로 일반 글자색으로 되돌아감)
    added = 0
    for std, parts in new_aliases.items():
        cell = desc_cells[std]
        current = _rich_text_to_str(cell.value).rstrip(", ").strip()
        added_text = ", ".join(parts)
        if current:
            cell.value = CellRichText([current, TextBlock(RED_FONT, ", " + added_text)])
        else:
            cell.value = CellRichText([TextBlock(RED_FONT, added_text)])
        added += len(parts)
        if log_fn: log_fn(f"  ✅ '{std}' 에 별칭 추가: '{added_text}'")

    if added > 0:
        wb.save(std_path)
        if log_fn: log_fn(f"표준불량명칭.xlsx 저장 완료 — {added}개 별칭 추가")
    return added


def norm(s):
    return re.sub(r'\s+', '', str(s)).lower()


def split_defect(raw):
    parts = re.split(r'\s*[/,+&]\s*', raw)
    return [p.strip() for p in parts if p.strip() and re.search(r'[가-힣a-zA-Z]', p)]


# ── 매핑 실행 ────────────────────────────────────────────────────
def build_mapping(raw_rows, std_names, adict, log_fn=None):
    slist  = [s[0] for s in std_names]
    catmap = {s[0]: s[1] for s in std_names}
    nalias = {norm(k): v for k, v in adict.items()}

    def map_one(part):
        part = part.strip()
        if not part: return None, 0, '-', False, ''
        if part in adict: return adict[part], 100, '정확일치', False, ''
        pn = norm(part)
        if pn in nalias: return nalias[pn], 100, '정확일치(정규화)', False, ''
        bs, bv = 0, None
        for k, v in adict.items():
            kn = norm(k)
            if len(kn) < 2: continue
            if kn in pn or pn in kn:
                sc = len(min(kn, pn, key=len)) / len(max(kn, pn, key=len)) * 100
                if sc > bs and sc >= TH_HIGH: bs, bv = sc, v
        if bv: return bv, int(bs), '별칭포함', False, ''
        nk = [n for n in slist if '기타' not in n]
        ki = [n for n in slist if '기타' in n]
        bn, bsc = None, 0
        for cands in [nk, ki]:
            for n in cands:
                sc = fuzz.token_sort_ratio(part, n)
                if sc > bsc: bsc, bn = sc, n
            if bsc >= TH_HIGH: break
        if bsc >= TH_HIGH:   return bn, bsc, '퍼지매핑', False, ''
        elif bsc >= TH_LOW:  return bn, bsc, '퍼지매핑', True, '수동검토필요'
        else:                return None, bsc, '미매핑', True, '적합한표준명없음'

    uraw  = list({r['defect_raw'] for r in raw_rows})
    cache = {}
    for i, raw in enumerate(uraw):
        if log_fn and i % 50 == 0:
            log_fn(f"매핑 중... {i}/{len(uraw)}")
        cache[raw] = [(p,) + map_one(p) for p in split_defect(raw)]
    if log_fn: log_fn(f"매핑 완료: {len(uraw)}개 불량명 처리")
    return cache, catmap

# ── 품목 유형 분류 ────────────────────────────────────────────────
_신발_KW = ['신발', 'SHOE', 'SHOES', 'SNEAKER', 'BOOT', 'BOOTS', 'SANDAL',
            '부츠', '샌들', '스니커', 'SLIPPER', '슬리퍼', 'LOAFER', 'HEEL']
_잡화_KW = ['가방', 'BAG', 'BAGS', '지갑', 'WALLET', '파우치', 'POUCH',
            '백팩', 'BACKPACK', '모자', 'HAT', 'CAP', '머플러', 'MUFFLER',
            'SCARF', '스카프', '장갑', 'GLOVE', 'GLOVES', '벨트', 'BELT',
            '우산', 'UMBRELLA', '숄더백', 'TOTE', '클러치', 'CLUTCH']

def classify_item_type(item_str: str) -> str:
    """품명(item) 필드로 의류/잡화/신발 판별. 기본값 의류."""
    if not item_str:
        return '의류'
    s = str(item_str).upper()
    for kw in _신발_KW:
        if kw in s:
            return '신발'
    for kw in _잡화_KW:
        if kw in s:
            return '잡화'
    return '의류'


def _load_sheet_names(wb, ptype: str):
    """잡화 파일에서 시트를 ptype(잡화/신발)에 맞게 선택."""
    for sn in wb.sheetnames:
        snu = sn.upper()
        if ptype == '신발' and '신발' in sn:
            return wb[sn]
        if ptype == '잡화' and ('가방' in sn or '지갑' in sn or '잡화' in sn):
            return wb[sn]
    return wb[wb.sheetnames[0]]


def load_standard_sheet(wb, ws) -> tuple:
    """워크시트 하나를 (names, adict) 로 파싱."""
    names, adict, cat = [], {}, None
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) < 3:
            continue
        c, no, name = row[0], row[1], row[2]
        desc = row[6] if len(row) > 6 else None
        desc_str = _rich_text_to_str(desc)
        if c:
            cat = c
        if not name:
            continue
        name = str(name).strip()
        if not name:
            continue
        names.append((name, cat, desc_str))
        adict[name] = name
        if desc_str:
            for cp in re.split(r'[,\n]', desc_str):
                for p in cp.split('/'):
                    p = p.strip()
                    if p and len(p) >= 2:
                        adict[p] = name
    return names, adict


def load_standard_typed(data_dir: str) -> dict:
    """
    data_dir 에서 의류/잡화/신발 표준불량명칭 로드.
    반환: {'의류': (names, adict), '잡화': (names, adict), '신발': (names, adict)}
    없는 파일은 기존 표준불량명칭.xlsx 로 폴백.
    """
    import os
    result = {}
    의류_path = os.path.join(data_dir, '[의류]표준불량명칭.xlsx')
    잡화_path = os.path.join(data_dir, '[잡화]표준불량명칭.xlsx')
    fallback   = os.path.join(data_dir, '..', '표준불량명칭.xlsx')

    # 의류
    if os.path.exists(의류_path):
        wb = openpyxl.load_workbook(의류_path, data_only=True)
        ws = wb[wb.sheetnames[0]]
        result['의류'] = load_standard_sheet(wb, ws)
    elif os.path.exists(fallback):
        wb = openpyxl.load_workbook(fallback, data_only=True)
        ws = find_std_sheet(wb)
        result['의류'] = load_standard_sheet(wb, ws)

    # 잡화 / 신발
    if os.path.exists(잡화_path):
        wb = openpyxl.load_workbook(잡화_path, data_only=True)
        for ptype in ['잡화', '신발']:
            ws = _load_sheet_names(wb, ptype)
            result[ptype] = load_standard_sheet(wb, ws)

    return result


def build_mapping_typed(raw_rows: list, std_by_type: dict, log_fn=None) -> tuple:
    """
    품목 유형(의류/잡화/신발)별로 다른 표준명칭을 사용해 매핑.
    raw_rows 각 행에 'product_type' 필드 추가(in-place).
    반환: (cache, catmap)  — 기존 build_mapping 과 동일한 인터페이스.
    """
    from collections import defaultdict

    # 각 행에 product_type 부여
    for r in raw_rows:
        r['product_type'] = classify_item_type(r.get('item', ''))

    # defect_raw → 대표 product_type (첫 등장 기준)
    raw_to_type = {}
    for r in raw_rows:
        if r['defect_raw'] not in raw_to_type:
            raw_to_type[r['defect_raw']] = r['product_type']

    # product_type 별로 고유 defect_raw 그룹화
    type_to_raws = defaultdict(list)
    for raw, ptype in raw_to_type.items():
        type_to_raws[ptype].append(raw)

    cache = {}
    catmap = {}
    fallback_type = '의류' if '의류' in std_by_type else next(iter(std_by_type))

    total = len(raw_to_type)
    done = 0

    for ptype, raws in type_to_raws.items():
        actual_type = ptype if ptype in std_by_type else fallback_type
        std_names, adict = std_by_type[actual_type]
        slist  = [s[0] for s in std_names]
        nalias = {norm(k): v for k, v in adict.items()}
        catmap.update({s[0]: s[1] for s in std_names})

        def map_one(part, _adict=adict, _nalias=nalias, _slist=slist):
            part = part.strip()
            if not part: return None, 0, '-', False, ''
            if part in _adict: return _adict[part], 100, '정확일치', False, ''
            pn = norm(part)
            if pn in _nalias: return _nalias[pn], 100, '정확일치(정규화)', False, ''
            bs, bv = 0, None
            for k, v in _adict.items():
                kn = norm(k)
                if len(kn) < 2: continue
                if kn in pn or pn in kn:
                    sc = len(min(kn, pn, key=len)) / len(max(kn, pn, key=len)) * 100
                    if sc > bs and sc >= TH_HIGH: bs, bv = sc, v
            if bv: return bv, int(bs), '별칭포함', False, ''
            nk = [n for n in _slist if '기타' not in n]
            ki = [n for n in _slist if '기타' in n]
            bn, bsc = None, 0
            for cands in [nk, ki]:
                for n in cands:
                    sc = fuzz.token_sort_ratio(part, n)
                    if sc > bsc: bsc, bn = sc, n
                if bsc >= TH_HIGH: break
            if bsc >= TH_HIGH:   return bn, bsc, '퍼지매핑', False, ''
            elif bsc >= TH_LOW:  return bn, bsc, '퍼지매핑', True, '수동검토필요'
            else:                return None, bsc, '미매핑', True, '적합한표준명없음'

        for raw in raws:
            cache[raw] = [(p,) + map_one(p) for p in split_defect(raw)]
            done += 1
            if log_fn and done % 50 == 0:
                log_fn(f"매핑 중... {done}/{total}")

    if log_fn: log_fn(f"매핑 완료: {total}개 불량명 처리")
    return cache, catmap



# ── 매핑 결과를 JSON 직렬화 가능한 리스트로 변환 ────────────────────
def mapping_to_records(raw_rows, cache, catmap):
    """API 응답용: 각 행을 dict로 변환"""
    records = []
    for r in raw_rows:
        for (part, std, sc, meth, rev, note) in cache.get(r['defect_raw'], []):
            records.append({
                **r,
                'part': part,
                'std': std or '',
                'score': sc,
                'method': meth,
                'review': rev,
                'note': note,
                'category': catmap.get(std, '') if std else '',
            })
    return records


# ── 매핑 통계 ────────────────────────────────────────────────────
def calc_stats(cache):
    all_items = [(std, sc, meth, rev)
                 for results in cache.values()
                 for (_, std, sc, meth, rev, _) in results]
    total = len(all_items)
    auto  = sum(1 for (_, sc, _, rev) in all_items if not rev and isinstance(sc, (int,float)) and sc >= TH_HIGH)
    unmap = sum(1 for (_, _, meth, _) in all_items if meth == '미매핑')
    review = total - auto - unmap
    pct = round(auto / total * 100, 1) if total else 0
    return {'total': total, 'auto': auto, 'review': review, 'unmapped': unmap, 'auto_pct': pct}


# ── Excel 출력 ───────────────────────────────────────────────────
def build_excel(raw_rows, cache, std_names, catmap, outpath, log_fn=None):
    G  = PatternFill("solid", fgColor="C6EFCE")
    Y  = PatternFill("solid", fgColor="FFEB9C")
    O  = PatternFill("solid", fgColor="FFB347")
    R  = PatternFill("solid", fgColor="FFC7CE")
    B  = PatternFill("solid", fgColor="4472C4")
    thin = Side(style='thin', color='BFBFBF')
    TBR  = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hdr(ws, r, c, t):
        x = ws.cell(r, c, t)
        x.font = Font(bold=True, color='FFFFFF')
        x.fill = B
        x.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        x.border = TBR

    def titrow(ws, t, n):
        ws.merge_cells('A1:' + ws.cell(1, n).column_letter + '1')
        x = ws['A1']
        x.value = t
        x.font = Font(bold=True, size=13, color='FFFFFF')
        x.fill = B
        x.alignment = Alignment(horizontal='center', vertical='center')

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Sheet 1
    if log_fn: log_fn("Sheet 1/4 — 불량명 매핑 사전 작성 중...")
    ws1 = wb.create_sheet('① 불량명 매핑 사전')
    h1 = ['분리 불량명', '원본 복합 표현', '표준 불량명(매핑)', '표준 카테고리', '신뢰도', '매핑 방법', '수동검토', '비고']
    titrow(ws1, '불량명 표준화 매핑 사전', len(h1))
    for ci, h in enumerate(h1, 1): hdr(ws1, 2, ci, h)
    ws1.row_dimensions[2].height = 28
    uraw = list({r['defect_raw'] for r in raw_rows})
    allr = []
    for raw in sorted(uraw):
        res = cache[raw]; ic = len(res) > 1; od = raw if ic else None
        for (part, std, sc, meth, rev, note) in res:
            cat = catmap.get(std, '') if std else ''
            n2  = note or ('복합' if ic else '')
            allr.append((part, od, std or '', cat, str(sc)+'%' if sc else '', meth, '요' if rev else '', n2, sc, std, rev))
    allr.sort(key=lambda x: x[0])
    for ri2, rd in enumerate(allr, 3):
        part, orig, std, cat, sc_s, meth, rev_s, note, sc_i, std_r, rev = rd
        if std_r is None: fill = R
        elif '기타' in (std_r or '') and meth not in ('정확일치','정확일치(정규화)','별칭포함'): fill = O
        elif isinstance(sc_i, (int, float)) and sc_i >= TH_HIGH: fill = G
        elif isinstance(sc_i, (int, float)) and sc_i >= TH_LOW:  fill = Y
        else: fill = R
        for ci, v in enumerate([part, orig, std, cat, sc_s, meth, rev_s, note], 1):
            x = ws1.cell(ri2, ci, v); x.fill = fill; x.border = TBR
            x.alignment = Alignment(vertical='center')
    for col, w in zip('ABCDEFGH', [28,42,20,28,10,18,10,20]):
        ws1.column_dimensions[col].width = w
    ws1.freeze_panes = 'A3'

    # Sheet 2
    if log_fn: log_fn("Sheet 2/4 — 표준명별 빈도 작성 중...")
    ws2 = wb.create_sheet('② 표준명별 빈도')
    h2 = ['카테고리', '표준 불량명', '불량 수량 합계', '발생 건수', '수량 비율(%)']
    titrow(ws2, '표준 불량명별 발생 빈도 요약', len(h2))
    for ci, h in enumerate(h2, 1): hdr(ws2, 2, ci, h)
    sq = defaultdict(int); sc2 = defaultdict(int)
    for r in raw_rows:
        for (p, std, sc, m, rv, n) in cache.get(r['defect_raw'], []):
            if std:
                sq[std] += r.get('qty_total', 0); sc2[std] += 1
    tq = sum(sq.values()) or 1
    for ri2, (sn, sc_, _) in enumerate(std_names, 3):
        for ci, v in enumerate([sc_, sn, sq.get(sn,0), sc2.get(sn,0), round(sq.get(sn,0)/tq*100, 2)], 1):
            ws2.cell(ri2, ci, v).border = TBR
    for col, w in zip('ABCDE', [30,22,16,12,14]):
        ws2.column_dimensions[col].width = w

    # Sheet 3
    if log_fn: log_fn("Sheet 3/4 — 미매핑 검토 목록 작성 중...")
    ws3 = wb.create_sheet('③ 미매핑 검토 목록')
    h3 = ['원본 불량명', '분리된 항목', '유사 표준명 후보 (top3)', '담당자 확정 표준명']
    titrow(ws3, '미매핑 / 수동검토 필요 항목', len(h3))
    for ci, h in enumerate(h3, 1): hdr(ws3, 2, ci, h)
    slist = [s[0] for s in std_names]
    ri3 = 3; seen = set()
    for raw in sorted(uraw):
        for (part, std, sc, meth, rev, note) in cache[raw]:
            if rev and (raw, part) not in seen:
                seen.add((raw, part))
                cands = sorted([(fuzz.token_sort_ratio(part, n), n) for n in slist], reverse=True)[:3]
                cs = ' / '.join([n+'('+str(s)+'%)' for s, n in cands])
                for ci, v in enumerate([raw, part, cs, std or ''], 1):
                    ws3.cell(ri3, ci, v).border = TBR
                ws3.cell(ri3, 4).fill = Y
                ri3 += 1
    for col, w in zip('ABCD', [42,28,60,22]):
        ws3.column_dimensions[col].width = w
    ws3.freeze_panes = 'A3'

    # Sheet 4
    if log_fn: log_fn("Sheet 4/4 — 불량상세 (표준명 적용) 작성 중...")
    ws4 = wb.create_sheet('④ 불량상세 (표준명 적용)')
    h4 = ['파일명','REPORT NO.','검사일','바이어','의뢰업체','브랜드','공장',
          '지역1','지역2','스타일번호','품명','검사수량(INSPEC)',
          '원본불량명','중불량','경불량','불량수량','분리 불량명','표준 불량명','신뢰도','매핑 방법']
    titrow(ws4, '불량상세 데이터 + 표준명 적용', len(h4))
    for ci, h in enumerate(h4, 1): hdr(ws4, 2, ci, h)
    dri = 3
    for r in raw_rows:
        for (part, std, sc, meth, rev, note) in cache.get(r['defect_raw'], []):
            vals = [r['file'], r['report_no'], r['date'], r['buyer'], r['client'], r['brand'],
                    r['factory'], r['region1'], r['region2'], r['style'], r['item'], r['inspec'],
                    r['defect_raw'], r['qty_mid'], r['qty_light'], r['qty_total'],
                    part, std or '', str(sc)+'%' if sc else '', meth]
            for ci, v in enumerate(vals, 1):
                ws4.cell(dri, ci, v).border = TBR
            dri += 1
    for ci, w in enumerate([18,18,12,14,14,12,14,10,12,16,16,10,32,8,8,8,28,20,8,14], 1):
        ws4.column_dimensions[ws4.cell(2, ci).column_letter].width = w
    ws4.freeze_panes = 'A3'

    if log_fn: log_fn("Excel 파일 저장 중...")
    wb.save(outpath)
    return dri - 3
