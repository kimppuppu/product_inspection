# -*- coding: utf-8 -*-
"""
performance_core.py — 제품평가팀 실적 분석 모듈
rawdata 형식: 접수일자 / 접수번호 / 공장명 / 검사종류 / 바이어 / 브랜드 /
              검사수량 / 합격수량 / 합격율 / 총수수료 / 발행금액(수수료포함) / 미발행금액 등

수익 기준:
  2026년    → 총수수료 (col 29)
  2023~2025 → 발행금액(수수료포함) (col 33)
"""
from __future__ import annotations
import re
from pathlib import Path
from collections import defaultdict
from datetime import datetime

try:
    import openpyxl
except ImportError:
    import subprocess, sys
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl


# ── 접수번호 파싱 ─────────────────────────────────────────────────
# 예: V152-26-03255  →  prefix=V, code=152, year=2026, seq=03255
_RNO_RE = re.compile(r'^([A-Za-z]+)(\d+)-(\d{2})-(\d+)')

def parse_report_no(rno: str) -> dict:
    rno = str(rno or '').strip()
    m = _RNO_RE.match(rno)
    if not m:
        return {'prefix': '', 'code': '', 'year': None, 'seq': ''}
    prefix = m.group(1).upper()
    code   = m.group(2)           # 131 or 152
    yr2    = int(m.group(3))
    year   = 2000 + yr2 if yr2 < 50 else 1900 + yr2
    return {'prefix': prefix, 'code': code, 'year': year, 'seq': m.group(4)}


# 접수번호 앞자리(prefix) → 국가/지역
#   S  중국 상해   Q  중국 청도   Y  중국 연태   A  인도네시아
#   V  베트남(특이사항에 '하노이'면 하노이, 그 외/공란이면 호치민·다낭)
#   H  한국(특이사항에 '미얀마'면 미얀마, 그 외/공란이면 한국)
#   L  대련 → 상해로 통합
REGION_ORDER = ['중국 상해', '중국 청도', '중국 연태', '인도네시아',
                '베트남 하노이', '베트남 호치민', '한국', '미얀마', '기타']

def get_region_label(prefix: str, note: str = '') -> str:
    """접수번호 앞자리(prefix) + 특이사항(note) → 8개 세분화 지역명"""
    p = (prefix or '').upper()
    note_l = (note or '').lower()
    if p in ('S', 'L'):   # L(대련)은 상해로 통합
        return '중국 상해'
    if p == 'Q':
        return '중국 청도'
    if p == 'Y':
        return '중국 연태'
    if p == 'A':
        return '인도네시아'
    if p == 'V':
        if '하노이' in (note or '') or 'hanoi' in note_l:
            return '베트남 하노이'
        return '베트남 호치민'   # 호치민/다낭/공란
    if p == 'H':
        if '미얀마' in (note or '') or 'myanmar' in note_l:
            return '미얀마'
        return '한국'           # 한국/공란
    return '기타'


# ── 수익 금액 선택 로직 ────────────────────────────────────────────
def get_revenue(row: dict) -> int:
    """연도별 수익 기준: 2026 → 총수수료, 나머지 → 발행금액(수수료포함)"""
    year = row.get('year')
    if year == 2026:
        return row.get('총수수료') or 0
    return row.get('발행금액_포함') or 0


# ── 숫자 변환 ─────────────────────────────────────────────────────
def _int(v):
    if v is None: return 0
    try: return int(str(v).replace(',', '').strip() or 0)
    except: return 0

def _float(v):
    if v is None: return 0.0
    try: return float(str(v).replace(',', '').replace('%', '').strip() or 0)
    except: return 0.0

def _ym(v) -> str | None:
    """날짜 → YYYY-MM (strptime 없이 직접 파싱 — 환경 의존성 없음)"""
    import re as _re
    if not v:
        return None
    # datetime/date 객체
    if hasattr(v, 'strftime'):
        return v.strftime('%Y-%m')
    # Excel 날짜 시리얼 정수 (예: 45678)
    if isinstance(v, (int, float)) and 30000 < v < 60000:
        try:
            from datetime import datetime as _dtt, timedelta as _td
            return (_dtt(1899, 12, 30) + _td(days=int(v))).strftime('%Y-%m')
        except Exception:
            pass
    s = str(v).strip()
    # YYYY-MM-DD, YYYY.MM.DD, YYYY/MM/DD
    m = _re.match(r'(\d{4})[-./](\d{1,2})', s)
    if m:
        y, mo = m.group(1), m.group(2).zfill(2)
        if 2000 <= int(y) <= 2099 and 1 <= int(mo) <= 12:
            return f"{y}-{mo}"
    # YYYYMMDD
    m = _re.match(r'(\d{4})(\d{2})(\d{2})', s)
    if m:
        y, mo = m.group(1), m.group(2)
        if 2000 <= int(y) <= 2099 and 1 <= int(mo) <= 12:
            return f"{y}-{mo}"
    return None


# ── 브랜드명 정규화/그룹핑 ────────────────────────────────────────
# 영구 참조 파일: "납품업체_브랜드_정리.xlsx" (영문 ↔ 국문 매핑표)
DEFAULT_BRAND_MAP_PATH = Path(__file__).resolve().parent.parent / "납품업체_브랜드_정리.xlsx"

# 마스터 파일에 없는 오타/축약형 별칭 (정규화 키 → 대표 브랜드명)
_EXTRA_ALIASES_RAW = {
    'NB':            'NEW BALANCE',
    'MLBKID':        'MLB KIDS',
    'HENRYCOTTON':   "HENRY COTTON'S",
    'HENRYCOTON':    "HENRY COTTON'S",
    'HERRYCOTTON':   "HENRY COTTON'S",
    'HERRYCOTON':    "HENRY COTTON'S",
    'DESCENTE':      'DESCENT',
    'DESCENTEGOLF':  'DESCENT GOLF',
    '아키이브앱크':   'ARCHIVEPKE',
    '젝시믹':        'XEXYMIX',
}

def _norm_key(s) -> str:
    """대소문자/공백/특수문자/캐리지리턴 제거한 비교용 키"""
    s = str(s or '').upper()
    s = s.replace('_X000D_', ' ')
    s = re.sub(r'[\r\n]', ' ', s)
    s = re.sub(r'[^A-Z0-9가-힣]', '', s)
    return s


def build_brand_alias_map(path: str | Path | None = None) -> dict:
    """
    납품업체_브랜드_정리.xlsx (영문/국문 매핑) + 수동 별칭표를 합쳐
    {정규화키: 대표브랜드명} 딕셔너리 생성.
    """
    path = Path(path) if path else DEFAULT_BRAND_MAP_PATH
    alias: dict[str, str] = {}
    try:
        if path.exists():
            wb = openpyxl.load_workbook(str(path), data_only=True)
            ws = wb[wb.sheetnames[0]]
            for ri, row in enumerate(ws.iter_rows(values_only=True)):
                if ri < 2:   # 헤더 2행 스킵
                    continue
                eng = str(row[1] or '').strip() if len(row) > 1 else ''
                kor = str(row[2] or '').strip() if len(row) > 2 else ''
                canonical = None
                if eng and eng != '-':
                    canonical = eng
                elif kor and kor != '-':
                    canonical = kor
                if not canonical:
                    continue
                for token in (eng, kor):
                    if token and token != '-':
                        k = _norm_key(token)
                        if k:
                            alias[k] = canonical
    except Exception:
        pass

    for raw_key, canonical in _EXTRA_ALIASES_RAW.items():
        alias[_norm_key(raw_key)] = canonical

    return alias


def normalize_brand(raw, alias_map: dict) -> str:
    """
    브랜드 원본 문자열 → 그룹핑된 대표 브랜드명.
    - "_x000D_", 줄바꿈, 앞뒤 공백/구분자 제거
    - "A/B/C" 같은 콤보 표기는 첫 항목 사용
    - "XEXYMIX(젝시믹스)", "젝시믹스(주)" 같은 괄호 표기 분해 후 매칭
    - 마스터/별칭표에 없으면 정리된 원본 그대로 반환 (그룹 자체는 유지)
    """
    s = str(raw or '')
    s = s.replace('_x000D_', ' ').replace('\r', ' ').replace('\n', ' ')
    s = s.strip(' /,;.\t')
    if not s or s in ('-', '.'):
        return '(미입력)'

    first = re.split(r'[\/,;]', s)[0].strip()
    if not first:
        first = s.strip()

    candidates = [first]
    m = re.match(r'^(.*?)\s*\((.*?)\)\s*$', first)
    if m:
        candidates = [m.group(1).strip(), m.group(2).strip()] + candidates

    for c in candidates:
        if not c:
            continue
        k = _norm_key(c)
        if k in alias_map:
            return alias_map[k]

    k = _norm_key(first)
    return alias_map.get(k, first)


# ── rawdata 로드 ──────────────────────────────────────────────────
SHEET_YEARS = ['2026', '2025', '2024', '2023']

CODE_LABELS = {
    '131': '원단검사',
    '152': '완제품 검사',
}

def _parse_sheet(ws, sheet_hint_year: int | None = None, log_fn=None,
                 brand_alias: dict | None = None) -> list[dict]:
    """
    단일 워크시트에서 행 파싱.
    sheet_hint_year: 연도 시트명에서 유추한 힌트 (None이면 접수일자/접수번호만 사용).
    """
    DATE_KEYS = ('접수일자','접수일','수령일자','수령일','날짜','일자','검사일','검사일자')
    RNO_KEYS  = ('접수번호','접수No','번호','검사번호')
    _MAP = {
        '공장명': 'factory', '검사종류': 'kind',
        '바이어': 'buyer', '브랜드': 'brand', '품명': 'item',
        '검사수량': 'inspec_qty', '합격수량': 'pass_qty',
        '합격율': 'pass_rate',
        '총수수료': 'fee_total',
        '저장수수료': 'fee_storage',
        '발행금액': 'fee_issued',
        '미발행금액': 'fee_unpaid',
        '검사비': 'fee_inspec',
        '출장비': 'fee_travel',
        '주요전달사항': 'note1',
        '바이어컨펌내역': 'note2',
        '특이사항': 'note3',
    }

    headers = None
    col: dict = {}
    rows: list[dict] = []
    brand_alias = brand_alias if brand_alias is not None else {}

    for raw in ws.iter_rows(values_only=True):
        # ── 헤더 탐지 ────────────────────────────────────────────
        if headers is None:
            row_strs = [str(c or '').strip() for c in raw]
            if any('접수' in s for s in row_strs) or (
                any('번호' in s for s in row_strs) and
                any('일자' in s or '날짜' in s for s in row_strs)
            ):
                headers = row_strs
                col = {}
                for ci, h in enumerate(headers):
                    if 'recv_date' not in col and any(k in h for k in DATE_KEYS):
                        col['recv_date'] = ci
                    if 'rno' not in col and any(k in h for k in RNO_KEYS):
                        col['rno'] = ci
                    if '발행금액' in h and '수수료' in h:
                        col['fee_issued_incl'] = ci
                    for key, alias in _MAP.items():
                        if key == h.strip():
                            col[alias] = ci
                if log_fn:
                    log_fn(f"    헤더: {headers[:5]}... 날짜열={col.get('recv_date','미발견')} 번호열={col.get('rno','미발견')}")
            continue

        if not any(raw):
            continue

        def g(alias, default=None):
            idx = col.get(alias)
            return raw[idx] if idx is not None and idx < len(raw) else default

        rno_val = str(g('rno') or '').strip()
        pinfo   = parse_report_no(rno_val)

        # ── 날짜 파싱 (접수일자 컬럼 우선, 실패 시 앞 컬럼 스캔) ──
        ym = _ym(g('recv_date'))
        if ym is None:
            for cv in list(raw)[:10]:
                c = _ym(cv)
                if c and (sheet_hint_year is None or int(c[:4]) == sheet_hint_year):
                    ym = c; break
        if ym is None and sheet_hint_year is not None:
            for cv in list(raw)[:5]:
                c = _ym(cv)
                if c and abs(int(c[:4]) - sheet_hint_year) <= 1:
                    ym = c; break

        # ── 연도 결정: 접수일자 → 접수번호 → 힌트 순 ──────────────
        year = (int(ym[:4]) if ym else None) or pinfo['year'] or sheet_hint_year
        if year is None:
            continue  # 연도 불명 → 스킵

        prefix = pinfo['prefix']
        code   = pinfo['code']
        # 지역 세분화: 주요전달사항 + 바이어컨펌내역 + 특이사항을 종합해 판별
        note = ' '.join(str(g(k) or '') for k in ('note1', 'note2', 'note3'))
        region_label = get_region_label(prefix, note)

        fee_total       = _int(g('fee_total'))
        fee_issued_incl = _int(raw[col['fee_issued_incl']]
                               if 'fee_issued_incl' in col and col['fee_issued_incl'] < len(raw)
                               else None)
        fee_unpaid      = _int(g('fee_unpaid'))

        brand_raw = str(g('brand') or '').strip()

        rows.append({
            'year':          year,
            'ym':            ym,
            'rno':           rno_val,
            'prefix':        prefix,
            'code':          code,
            'region_label':  region_label,
            'factory':       str(g('factory') or '').strip(),
            'kind':          str(g('kind')    or '').strip(),
            'buyer':         str(g('buyer')   or '').strip(),
            'brand':         normalize_brand(brand_raw, brand_alias),
            'brand_raw':     brand_raw,
            'item':          str(g('item')    or '').strip(),
            'inspec_qty':    _int(g('inspec_qty')),
            'pass_qty':      _int(g('pass_qty')),
            'pass_rate':     _float(g('pass_rate')),
            '총수수료':       fee_total,
            '발행금액_포함':  fee_issued_incl,
            '미발행금액':     fee_unpaid,
        })

    return rows


# 통합 시트로 인식할 이름 패턴 (정확히 일치 또는 포함)
_UNIFIED_SHEET_NAMES = {'전체', 'rawdata', 'raw', 'raw data', '전체데이터', 'data', '통합'}


def _is_unified_sheet(name: str) -> bool:
    n = name.strip().lower()
    if n in _UNIFIED_SHEET_NAMES:
        return True
    return any(key in n for key in ('rawdata', 'raw data', '전체', '통합'))


def load_performance(path: str, log_fn=None, brand_map_path: str | Path | None = None) -> list[dict]:
    """
    엑셀 파일 로드 — 두 가지 구조 자동 인식:
      ① 통합 시트 ('전체', 'rawdata(2401~2605)' 등): 모든 연도가 하나의 시트에
      ② 연도별 시트 ('2023','2024','2025','2026'): 기존 방식
    year는 접수일자 → 접수번호 → 시트명 순으로 결정.
    브랜드명은 납품업체_브랜드_정리.xlsx 기준으로 정규화/그룹핑됨.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    rows: list[dict] = []
    brand_alias = build_brand_alias_map(brand_map_path)
    if log_fn:
        log_fn(f"  브랜드 그룹핑 매핑 {len(brand_alias):,}건 로드")

    # ── ① 통합 시트 우선 확인 (이름에 'rawdata'/'전체'/'통합' 포함) ──
    unified = next((sn for sn in wb.sheetnames if _is_unified_sheet(sn)), None)
    if unified:
        if log_fn: log_fn(f"  통합 시트 '{unified}' 읽는 중...")
        rows = _parse_sheet(wb[unified], sheet_hint_year=None, log_fn=log_fn, brand_alias=brand_alias)
        if log_fn: log_fn(f"  총 {len(rows):,}건 로드 완료")
        return rows

    # ── ② 연도별 시트 ────────────────────────────────────────────
    for sn in wb.sheetnames:
        if sn not in SHEET_YEARS:
            continue
        if log_fn: log_fn(f"  {sn}년 시트 읽는 중...")
        sheet_rows = _parse_sheet(wb[sn], sheet_hint_year=int(sn), log_fn=log_fn, brand_alias=brand_alias)
        rows.extend(sheet_rows)
        if log_fn: log_fn(f"    → {len(sheet_rows):,}건")

    if log_fn: log_fn(f"  총 {len(rows):,}건 로드 완료")
    return rows


def create_unified_excel(src_path: str, dst_path: str, log_fn=None) -> str:
    """
    연도별 시트 → 단일 '전체' 시트로 통합한 새 엑셀 파일 생성.
    반환값: dst_path
    """
    wb_src = openpyxl.load_workbook(src_path, data_only=True)

    # 헤더 추출 (첫 번째 연도 시트에서)
    header_row = None
    data_sheets = [sn for sn in wb_src.sheetnames if sn in SHEET_YEARS]
    for sn in data_sheets:
        ws = wb_src[sn]
        for raw in ws.iter_rows(values_only=True):
            row_strs = [str(c or '').strip() for c in raw]
            if any('접수' in s for s in row_strs):
                header_row = list(raw)
                break
        if header_row:
            break

    if not header_row:
        raise ValueError("헤더 행을 찾을 수 없습니다.")

    wb_dst = openpyxl.Workbook()
    ws_dst = wb_dst.active
    ws_dst.title = '전체'
    ws_dst.append(header_row)

    total = 0
    for sn in data_sheets:
        ws_src = wb_src[sn]
        header_found = False
        for raw in ws_src.iter_rows(values_only=True):
            if not header_found:
                row_strs = [str(c or '').strip() for c in raw]
                if any('접수' in s for s in row_strs):
                    header_found = True
                continue
            if any(v is not None for v in raw):
                ws_dst.append(list(raw))
                total += 1
        if log_fn: log_fn(f"  {sn}년 시트 복사 완료")

    wb_dst.save(dst_path)
    if log_fn: log_fn(f"  통합 파일 저장: {dst_path} ({total:,}건)")
    return dst_path


# ── 수익 계산 헬퍼 ────────────────────────────────────────────────
def revenue(row: dict) -> int:
    return row['총수수료'] if row['year'] == 2026 else row['발행금액_포함']


# ── 필터 ─────────────────────────────────────────────────────────
def filter_rows(rows: list[dict], years: list[int] = None,
                region: str = None, code: str = None,
                buyer: str = None) -> list[dict]:
    out = []
    for r in rows:
        if years and r['year'] not in years: continue
        if region and r['region_label'] != region: continue
        if code   and r['code'] != code:           continue
        if buyer  and r['buyer'] != buyer:          continue
        out.append(r)
    return out


# ── 집계 함수들 ───────────────────────────────────────────────────

def summary_by_year(rows: list[dict]) -> list[dict]:
    """연도별 합계"""
    agg = defaultdict(lambda: {'cnt': 0, 'inspec': 0, 'revenue': 0, 'unpaid': 0})
    for r in rows:
        y = r['year']
        agg[y]['cnt']     += 1
        agg[y]['inspec']  += r['inspec_qty']
        agg[y]['revenue'] += revenue(r)
        agg[y]['unpaid']  += r['미발행금액']
    return [{'year': y, **v} for y, v in sorted(agg.items())]


def summary_by_month(rows: list[dict]) -> list[dict]:
    """월별 건수·수익 (연도 포함)"""
    agg = defaultdict(lambda: {'cnt': 0, 'inspec': 0, 'revenue': 0})
    for r in rows:
        ym = r['ym'] or f"{r['year']}-??"
        agg[ym]['cnt']     += 1
        agg[ym]['inspec']  += r['inspec_qty']
        agg[ym]['revenue'] += revenue(r)
    return [{'ym': ym, **v} for ym, v in sorted(agg.items())]


def summary_by_region(rows: list[dict]) -> list[dict]:
    """지역별 집계 (8개 권역 순서대로 정렬)"""
    agg = defaultdict(lambda: {'cnt': 0, 'inspec': 0, 'revenue': 0, 'unpaid': 0})
    for r in rows:
        reg = r['region_label'] or '기타'
        agg[reg]['cnt']     += 1
        agg[reg]['inspec']  += r['inspec_qty']
        agg[reg]['revenue'] += revenue(r)
        agg[reg]['unpaid']  += r['미발행금액']

    def _order(reg):
        return REGION_ORDER.index(reg) if reg in REGION_ORDER else len(REGION_ORDER)

    return sorted(
        [{'region': reg, **v} for reg, v in agg.items()],
        key=lambda x: _order(x['region'])
    )


def summary_by_buyer(rows: list[dict], top_n: int = 20) -> list[dict]:
    """바이어별 집계 (수익 TOP N)"""
    agg = defaultdict(lambda: {'cnt': 0, 'inspec': 0, 'revenue': 0})
    for r in rows:
        b = r['buyer'] or '(미입력)'
        agg[b]['cnt']     += 1
        agg[b]['inspec']  += r['inspec_qty']
        agg[b]['revenue'] += revenue(r)
    result = sorted(
        [{'buyer': b, **v} for b, v in agg.items()],
        key=lambda x: -x['revenue']
    )
    return result[:top_n]


def summary_by_brand(rows: list[dict], top_n: int = 20) -> list[dict]:
    """브랜드별 집계 (그룹핑 후 수익 TOP N)"""
    agg = defaultdict(lambda: {'cnt': 0, 'inspec': 0, 'revenue': 0})
    for r in rows:
        b = r['brand'] or '(미입력)'
        agg[b]['cnt']     += 1
        agg[b]['inspec']  += r['inspec_qty']
        agg[b]['revenue'] += revenue(r)
    result = sorted(
        [{'brand': b, **v} for b, v in agg.items()],
        key=lambda x: -x['revenue']
    )
    return result[:top_n]


def summary_by_code(rows: list[dict]) -> list[dict]:
    """코드별(131/152) 집계"""
    agg = defaultdict(lambda: {'cnt': 0, 'inspec': 0, 'revenue': 0})
    for r in rows:
        c = r['code'] or '기타'
        agg[c]['cnt']     += 1
        agg[c]['inspec']  += r['inspec_qty']
        agg[c]['revenue'] += revenue(r)
    return sorted(
        [{'code': c, 'label': CODE_LABELS.get(c, c), **v} for c, v in agg.items()],
        key=lambda x: -x['revenue']
    )


def region_code_crosstab(rows: list[dict]) -> list[dict]:
    """지역(8개 권역) × 코드(131 원단검사 / 152 완제품검사) 교차 집계"""
    agg: dict[str, dict[str, int]] = defaultdict(lambda: defaultdict(int))
    for r in rows:
        reg = r['region_label'] or '기타'
        c = r['code'] if r['code'] in CODE_LABELS else '기타'
        agg[reg][c] += revenue(r)

    def _order(reg):
        return REGION_ORDER.index(reg) if reg in REGION_ORDER else len(REGION_ORDER)

    out = []
    for reg, d in sorted(agg.items(), key=lambda kv: _order(kv[0])):
        out.append({
            'region': reg,
            'c131':  d.get('131', 0),
            'c152':  d.get('152', 0),
            'other': sum(v for c, v in d.items() if c not in ('131', '152')),
            'total': sum(d.values()),
        })
    return out


def yoy_comparison(rows: list[dict], dim: str | None = 'region',
                   same_months: list[str] | None = None,
                   top_n: int | None = None,
                   sort_year: int = 2025) -> list[dict]:
    """
    연도별(2023~2026) 비교 — 동기누적 비교 시 same_months 지정.
    dim: 'region' | 'buyer' | 'brand' | 'code' | None(전체 단일 그룹)
    same_months: ['01','02',...] — 지정 시 해당 월만 집계 (동기대비 누적)
    top_n: 결과 개수 제한 (sort_year 기준 정렬 후 상위 N개)
    반환: [{label, y2023, y2024, y2025, y2026, growth_24_25, growth_25_26}]
    """
    if dim is None:
        key_fn = lambda r: '전체'
    else:
        key_fn = {
            'region': lambda r: r['region_label'] or '기타',
            'buyer':  lambda r: r['buyer'] or '(미입력)',
            'brand':  lambda r: r['brand'] or '(미입력)',
            'code':   lambda r: CODE_LABELS.get(r['code'], r['code'] or '기타'),
        }[dim]

    agg: dict[str, dict[int, int]] = defaultdict(lambda: defaultdict(int))
    cnt: dict[str, dict[int, int]] = defaultdict(lambda: defaultdict(int))
    for r in rows:
        # 동기대비: 지정 월만 집계 (연도 구분 없이 ym이 있는 행 기준)
        if same_months is not None:
            mm = (r['ym'] or '')[-2:] if r['ym'] else None
            if mm not in same_months:
                continue
        key = key_fn(r) or '기타'
        agg[key][r['year']] += revenue(r)
        cnt[key][r['year']] += 1

    years = [2023, 2024, 2025, 2026]
    result = []
    for key, yr_data in agg.items():
        row = {'label': key}
        for y in years:
            row[f'y{y}']  = yr_data.get(y, 0)
            row[f'c{y}']  = cnt[key].get(y, 0)
        prev24 = yr_data.get(2024, 0)
        prev25 = yr_data.get(2025, 0)
        curr26 = yr_data.get(2026, 0)
        row['growth_24_25'] = round((prev25 - prev24) / prev24 * 100, 1) if prev24 else None
        row['growth_25_26'] = round((curr26 - prev25) / prev25 * 100, 1) if prev25 else None
        row['growth'] = row['growth_25_26']  # 하위호환
        result.append(row)

    result.sort(key=lambda x: -x.get(f'y{sort_year}', 0))
    if top_n:
        result = result[:top_n]
    return result


def cumulative_by_year(rows: list[dict], same_months: list[str] | None,
                        years: tuple[int, ...] = (2024, 2025, 2026)) -> dict:
    """
    동기누적(예: 1~5월) 실적 — 연도별 합계.
    latest(=max(years))는 same_months 제한 없이 해당 연도 전체(이미 1~5월까지만 존재)를 집계,
    그 외 연도는 same_months에 해당하는 월만 집계.
    반환: {year: {rev, cnt}}
    """
    latest = max(years)
    out = {}
    for y in years:
        if y == latest:
            sub = [r for r in rows if r['year'] == y]
        else:
            sub = [
                r for r in rows
                if r['year'] == y and r.get('ym') and len(r['ym']) == 7
                and (same_months is None or r['ym'][-2:] in same_months)
            ]
        out[y] = {
            'rev': sum(get_revenue(r) for r in sub),
            'cnt': len(sub),
        }
    return out


def monthly_compare(rows: list[dict], dim: str | None = None,
                     group_filter: set | list | None = None,
                     years: tuple[int, ...] = (2024, 2025, 2026)) -> dict:
    """
    그룹별 × 연도별 × 월별 실적 (3개년 월별 비교용).
    dim: None(전체 단일 그룹='전체') | 'region' | 'buyer' | 'brand' | 'code'
    group_filter: 포함할 그룹 라벨 집합 (예: TOP20 바이어/브랜드 라벨). None이면 전체 그룹.
    반환: { group_label: { year: {mm: revenue} } }
    """
    if dim is None:
        key_fn = lambda r: '전체'
    else:
        key_fn = {
            'region': lambda r: r['region_label'] or '기타',
            'buyer':  lambda r: r['buyer'] or '(미입력)',
            'brand':  lambda r: r['brand'] or '(미입력)',
            'code':   lambda r: CODE_LABELS.get(r['code'], r['code'] or '기타'),
        }[dim]

    out: dict[str, dict[int, dict[str, int]]] = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))
    for r in rows:
        if r['year'] not in years:
            continue
        if not r.get('ym') or len(r['ym']) != 7:
            continue
        g = key_fn(r) or '기타'
        if group_filter is not None and g not in group_filter:
            continue
        mm = r['ym'][-2:]
        out[g][r['year']][mm] += revenue(r)

    return {g: {y: dict(mms) for y, mms in yd.items()} for g, yd in out.items()}


def actual_by_month_code(rows: list[dict], year: int) -> dict:
    """지정 연도의 월별 × 코드(131/152) 실적 — 목표대비용"""
    out: dict[str, dict] = {}
    for r in rows:
        if r['year'] != year or not r.get('ym') or len(r['ym']) != 7:
            continue
        mm = r['ym'][-2:]
        d = out.setdefault(mm, {'131': 0, '152': 0, 'total': 0})
        rev = revenue(r)
        d['total'] += rev
        if r['code'] in ('131', '152'):
            d[r['code']] += rev
    return out


def load_plan_budget(path: str) -> dict:
    """
    수입계획(예산) 엑셀 파일 로드. ('주관 실적' 시트, 단위: 천원)
    헤더행에서 '1월'이 두 번 나오면 첫 번째=국내계산서, 두 번째=해외계산서로 인식해 합산.
    코드(131/152) 행은 '세분류(코드)' 등 '코드'가 포함된 헤더 열에서 식별.
    반환: {annual: {131, 152, total}, monthly: {"01":{131,152,total}, ...}}
    값 단위: 원 (파일 원본이 천원이면 *1000 적용)
    """
    wb = openpyxl.load_workbook(path, data_only=True)

    target_sheet = None
    for sn in wb.sheetnames:
        if '주관' in sn or '실적' in sn or '계획' in sn or 'budget' in sn.lower():
            target_sheet = sn
            break
    if not target_sheet:
        target_sheet = wb.sheetnames[0]
    ws = wb[target_sheet]

    months_mm = ['01','02','03','04','05','06','07','08','09','10','11','12']

    code_col = None
    domestic_start = None
    overseas_start = None

    all_rows = list(ws.iter_rows(values_only=True))
    header_ri = None

    for ri, row in enumerate(all_rows):
        row_strs = [str(c or '').strip() for c in row]
        # '1월' 셀이 있는 행이 헤더 (국내/해외 두 번 등장 가능)
        jan_idx = [ci for ci, s in enumerate(row_strs) if s == '1월']
        if jan_idx:
            header_ri = ri
            domestic_start = jan_idx[0]
            overseas_start = jan_idx[1] if len(jan_idx) > 1 else None
            # 코드 컬럼: '코드'가 포함된 헤더
            for ci, s in enumerate(row_strs):
                if '코드' in s:
                    code_col = ci
                    break
            if code_col is None:
                for ci, s in enumerate(row_strs[:8]):
                    if s.strip():
                        code_col = ci
                        break
            break

    if header_ri is None or domestic_start is None:
        return {'annual': {'131': 0, '152': 0, 'total': 0}, 'monthly': {}}

    def _read_monthly(row):
        result = {}
        for mi in range(12):
            ci_d = domestic_start + mi
            ci_o = (overseas_start + mi) if overseas_start is not None else None
            d = _int(row[ci_d]) if ci_d < len(row) else 0
            o = (_int(row[ci_o]) if ci_o < len(row) else 0) if ci_o is not None else 0
            result[months_mm[mi]] = (d + o) * 1000  # 천원 → 원
        return result

    plan_131: dict = {}
    plan_152: dict = {}

    for row in all_rows[header_ri + 1:]:
        if not row or all(v is None for v in row):
            continue
        code_val = str(row[code_col] if code_col is not None and code_col < len(row) else '').strip()
        if '131' in code_val and not plan_131:
            plan_131 = _read_monthly(row)
        elif '152' in code_val and not plan_152:
            plan_152 = _read_monthly(row)

    monthly = {}
    for mm in months_mm:
        v131 = plan_131.get(mm, 0)
        v152 = plan_152.get(mm, 0)
        monthly[mm] = {'131': v131, '152': v152, 'total': v131 + v152}

    annual = {
        '131':  sum(plan_131.values()),
        '152':  sum(plan_152.values()),
        'total': sum(plan_131.values()) + sum(plan_152.values()),
    }
    return {'annual': annual, 'monthly': monthly}


def get_filter_options(rows: list[dict]) -> dict:
    years   = sorted({r['year'] for r in rows if r['year']})
    regions = [r for r in REGION_ORDER if r in {x['region_label'] for x in rows if x['region_label']}]
    codes   = sorted({r['code'] for r in rows if r['code']})
    buyers  = sorted({r['buyer'] for r in rows if r['buyer']})[:50]
    return {'years': years, 'regions': regions, 'codes': codes, 'buyers': buyers}
